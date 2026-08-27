#!/usr/bin/env python3
"""
Consolida as abas Mizuno / Olympikus / Under Armour de um export da base HANA
em uma única aba "Base Final", seguindo o mapeamento fixo descrito no SKILL.md.

Uso:
    python consolidate.py <arquivo_entrada.xlsx> [<arquivo_saida.xlsx>]

Se o arquivo de saída não for informado, é gerado como
"<nome_original>_Final.xlsx" na mesma pasta do arquivo de entrada.
"""

import sys
import unicodedata
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuração do mapeamento (edite aqui se a estrutura da base HANA mudar)
# ---------------------------------------------------------------------------

# Nome esperado de cada aba de origem -> chaves usadas para localizá-la mesmo
# com variações de grafia/maiúsculas (comparação normalizada, ver _norm()).
SOURCE_SHEETS = {
    "Mizuno": ["mizuno"],
    "Olympikus": ["olympikus", "olympicus"],
    "Under Armour": ["under armour", "underarmour", "under_armour", "ua"],
}

# Cabeçalhos que buscamos em cada aba, por NOME (não por letra de coluna),
# porque a posição das colunas varia entre abas (ex.: Under Armour vem
# deslocada 1 coluna para a direita em relação a Mizuno/Olympikus).
# Cada entrada é uma lista de possíveis grafias normalizadas do cabeçalho;
# a primeira que bater na linha 1 da aba é usada.
HEADER_ALIASES = {
    "sku": ["sku"],
    "codigo_barras": ["codigo de barras", "codigodebarras", "cod barras"],
    "material_vulca_tam": ["materialvulcatam"],
    "material_vulca": ["materialvulca"],
    "artigo": ["artigo"],
    "descricao_item": ["descricao do item"],
    "nome_grupo": ["nome do grupo"],
    "colorway": ["colorway description", "colorway descri"],  # aceita truncado
}

# Ordem e nomes das colunas na aba final. Cada tupla é (nome_da_coluna_final,
# chave_do_header_acima). "marca" e "segmento" são calculadas, não vêm direto
# de um header.
FINAL_COLUMNS = [
    ("SKU", "sku"),
    ("Códigodebarras", "codigo_barras"),
    ("MaterialVulcaTam", "material_vulca_tam"),
    ("MaterialVulca", "material_vulca"),
    ("Artigo", "artigo"),
    ("SKU", "sku"),
    ("Descrição do item", "descricao_item"),
    ("Nome do grupo", "nome_grupo"),
    ("Códigodebarras", "codigo_barras"),
    ("Colorway Description", "colorway"),
    ("MARCA", None),      # calculada: nome da aba de origem
    ("SEGMENTO", None),   # calculada: a partir de "nome do grupo"
]

# Marcas que não têm aba própria: vivem dentro da aba de outra marca e são
# identificadas apenas pelo texto de "Nome do grupo". Ex.: OPANKA não tem
# estoque próprio, mas tem cadastro na base de embalagem, e vem dentro da aba
# Olympikus. Formato: (substring procurada em "Nome do grupo", MARCA, SEGMENTO
# fixo ou None para cair nas regras normais).
MARCA_EMBUTIDA_RULES = [
    ("opanka", "OPANKA", "CHINELOS"),
]

# Regras que dependem do texto exato de "Nome do grupo" e têm precedência
# sobre as regras genéricas de segmento abaixo. Ex.: a linha de roupas do
# Botafogo da Mizuno vem marcada como "FUTEBOL MIZUNO", mas é vestuário —
# a palavra "futebol" sozinha não diria isso.
SEGMENTO_OVERRIDES = [
    ("futebol mizuno", "VESTUÁRIOS"),
]

# Regras de tradução/padronização do SEGMENTO a partir do texto de
# "Nome do grupo" (comparação por substring, case/acento-insensitive).
# A ordem importa: a primeira regra que bater é usada — por isso "chinelos" e
# "meias" vêm antes de "calcados"/"vestuario", já que um grupo pode conter
# as duas palavras e o mais específico deve ganhar.
SEGMENTO_RULES = [
    ("chinelos", "CHINELOS"),
    ("chinelo", "CHINELOS"),
    ("sandals", "CHINELOS"),
    ("slides", "CHINELOS"),
    ("meias", "MEIAS"),
    ("meia", "MEIAS"),
    ("socks", "MEIAS"),
    ("calcados", "CALÇADOS"),
    ("footwear", "CALÇADOS"),
    ("acessorios", "ACESSÓRIOS"),
    ("accessories", "ACESSÓRIOS"),
    ("apparel", "VESTUÁRIOS"),
    ("vestuario", "VESTUÁRIOS"),
    ("vestuarios", "VESTUÁRIOS"),
]

FINAL_SHEET_NAME = "Base Final"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _norm(text) -> str:
    """Normaliza texto para comparação: sem acento, minúsculo, sem espaços extras."""
    if text is None:
        return ""
    text = str(text).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return " ".join(text.split())


def find_sheet(wb, aliases):
    """Encontra a primeira aba do workbook cujo nome normalizado bate com um dos aliases."""
    for sheet_name in wb.sheetnames:
        norm_name = _norm(sheet_name)
        for alias in aliases:
            if _norm(alias) in norm_name or norm_name in _norm(alias):
                return sheet_name
    return None


def map_headers(header_row):
    """
    header_row: lista de valores da linha 1 da aba (na ordem das colunas).
    Retorna {header_key: column_index (0-based)} para os headers que foram encontrados.
    """
    # Correspondência EXATA (após normalizar acento/caixa) — importante para não
    # confundir "SKU" com uma coluna parecida como "SKUCOMERCIAL?".
    found = {}
    normalized = [_norm(v) for v in header_row]
    for key, aliases in HEADER_ALIASES.items():
        alias_norms = {_norm(a) for a in aliases}
        for col_idx, cell_norm in enumerate(normalized):
            if cell_norm in alias_norms:
                found[key] = col_idx
                break
    return found


def compute_marca_segmento(nome_grupo_valor, marca_da_aba):
    """
    Decide MARCA e SEGMENTO de uma linha a partir do texto de "Nome do grupo".

    A MARCA normalmente é a aba de origem, mas algumas marcas não têm aba
    própria e só se distinguem pelo "Nome do grupo" (ver MARCA_EMBUTIDA_RULES),
    então essa checagem vem primeiro.

    Retorna (marca, segmento, segmento_foi_mapeado).
    """
    norm = _norm(nome_grupo_valor)

    # 1) Marca embutida na aba de outra marca (ex.: OPANKA dentro de Olympikus).
    for needle, marca, segmento_fixo in MARCA_EMBUTIDA_RULES:
        if needle in norm:
            if segmento_fixo:
                return marca, segmento_fixo, True
            marca_da_aba = marca
            break

    # 2) Overrides de segmento por grupo específico (ex.: FUTEBOL MIZUNO).
    for needle, segmento in SEGMENTO_OVERRIDES:
        if needle in norm:
            return marca_da_aba, segmento, True

    # 3) Regras genéricas de tradução/padronização.
    for needle, segmento in SEGMENTO_RULES:
        if needle in norm:
            return marca_da_aba, segmento, True

    # Nada bateu: mantém o texto original para revisão manual.
    original = str(nome_grupo_valor).strip() if nome_grupo_valor else ""
    return marca_da_aba, original, False


# ---------------------------------------------------------------------------
# Núcleo da consolidação
# ---------------------------------------------------------------------------

def consolidate(input_path: Path):
    wb = load_workbook(input_path, data_only=True)

    report_lines = []
    all_rows = []  # linhas já no formato final (lista de valores)
    missing_field_counts = {}  # marca -> quantidade de linhas com campo faltando
    unmapped_segmentos = set()
    marca_counts = {}  # MARCA final -> quantidade de linhas (pode diferir da aba,
                       # ex.: OPANKA sai de dentro da aba Olympikus)

    for marca, aliases in SOURCE_SHEETS.items():
        sheet_name = find_sheet(wb, aliases)
        if sheet_name is None:
            report_lines.append(f"⚠️  Aba '{marca}' não encontrada no arquivo — pulando.")
            continue

        ws = wb[sheet_name]
        header_row = [cell.value for cell in ws[1]]
        header_map = map_headers(header_row)

        missing_headers = [k for k in HEADER_ALIASES if k not in header_map]
        if missing_headers:
            report_lines.append(
                f"⚠️  Aba '{sheet_name}': cabeçalhos não encontrados: {', '.join(missing_headers)} "
                f"(essas colunas ficarão em branco na base final para esta aba)."
            )

        rows_in_sheet = 0
        missing_in_sheet = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue  # linha totalmente vazia

            def get(key):
                idx = header_map.get(key)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            nome_grupo_valor = get("nome_grupo")
            marca_linha, segmento, mapped_ok = compute_marca_segmento(nome_grupo_valor, marca)
            if not mapped_ok and nome_grupo_valor:
                unmapped_segmentos.add(str(nome_grupo_valor).strip())
            marca_counts[marca_linha] = marca_counts.get(marca_linha, 0) + 1

            final_row = []
            row_missing = False
            for col_name, header_key in FINAL_COLUMNS:
                if header_key == "sku":
                    val = get("sku")
                elif header_key == "codigo_barras":
                    val = get("codigo_barras")
                elif header_key is not None:
                    val = get(header_key)
                elif col_name == "MARCA":
                    val = marca_linha
                elif col_name == "SEGMENTO":
                    val = segmento
                else:
                    val = None
                if header_key is not None and (val is None or str(val).strip() == ""):
                    row_missing = True
                final_row.append(val)

            all_rows.append(final_row)
            rows_in_sheet += 1
            if row_missing:
                missing_in_sheet += 1

        missing_field_counts[marca] = missing_in_sheet
        report_lines.append(f"✓ Aba '{sheet_name}': {rows_in_sheet} linhas consolidadas.")

    if marca_counts:
        resumo_marcas = ", ".join(f"{m}: {c}" for m, c in sorted(marca_counts.items()))
        report_lines.append(f"Linhas por MARCA na base final — {resumo_marcas}")

    if unmapped_segmentos:
        report_lines.append(
            "⚠️  Valores de 'Nome do grupo' sem regra de SEGMENTO mapeada (mantidos como estavam, "
            "revisar manualmente): " + "; ".join(sorted(unmapped_segmentos))
        )

    for marca, count in missing_field_counts.items():
        if count:
            report_lines.append(f"⚠️  Aba '{marca}': {count} linha(s) com pelo menos um campo ausente.")

    return all_rows, report_lines


def write_output(rows, output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = FINAL_SHEET_NAME

    headers = [name for name, _ in FINAL_COLUMNS]
    ws.append(headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        ws.append(row)

    # Auto-largura simples baseada no maior valor de cada coluna.
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            v = row[col_idx - 1]
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    ws.freeze_panes = "A2"
    wb.save(output_path)


def main():
    if len(sys.argv) < 2:
        print("Uso: python consolidate.py <arquivo_entrada.xlsx> [<arquivo_saida.xlsx>]")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    if not input_path.exists():
        print(f"Arquivo não encontrado: {input_path}")
        sys.exit(1)

    if len(sys.argv) >= 3:
        output_path = Path(sys.argv[2])
    else:
        output_path = input_path.with_name(f"{input_path.stem}_Final.xlsx")

    rows, report_lines = consolidate(input_path)
    write_output(rows, output_path)

    print(f"\nBase final gerada: {output_path}")
    print(f"Total de linhas consolidadas: {len(rows)}\n")
    print("Resumo:")
    for line in report_lines:
        print(f"  {line}")


if __name__ == "__main__":
    main()
