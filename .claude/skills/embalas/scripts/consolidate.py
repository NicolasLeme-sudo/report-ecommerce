#!/usr/bin/env python3
"""
Consolida as abas Mizuno / Olympikus / Under Armour de um export da base HANA
em uma única aba "Base Final", seguindo o mapeamento fixo descrito no SKILL.md.

Uso:
    python consolidate.py <arquivo1> [<arquivo2> <arquivo3> ...] [--output <arquivo_saida.xlsx>]

Aceita tanto UM arquivo com as 3 abas (Mizuno/Olympikus/Under Armour) quanto
até 3 arquivos separados, um por marca (nesse caso a aba pode ter qualquer
nome — a marca é inferida pelo nome do arquivo). Lê .xlsx e .xlsb.

Se --output não for informado, é gerado "Embalas.xlsx" na mesma pasta do
primeiro arquivo de entrada — esse é sempre o nome/formato de saída padrão.
"""

import sys
import unicodedata
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import pyxlsb
except ImportError:
    pyxlsb = None

# ---------------------------------------------------------------------------
# Configuração do mapeamento (edite aqui se a estrutura da base HANA mudar)
# ---------------------------------------------------------------------------

# Nome de cada marca de origem -> aliases usados para localizá-la, tanto pelo
# nome da ABA (workbook único com 3 abas) quanto pelo nome do ARQUIVO (quando
# cada marca chega em um arquivo separado, o que aconteceu na prática quando
# o export ficou pesado demais para mandar em um workbook só).
SOURCE_SHEETS = {
    "Mizuno": ["mizuno"],
    "Olympikus": ["olympikus", "olympicus"],
    "Under Armour": ["under armour", "underarmour", "under_armour", "ua"],
}

# Cabeçalhos que buscamos em cada aba, por NOME (não por letra de coluna),
# porque a posição das colunas varia entre abas (ex.: Under Armour vem
# deslocada 1 coluna para a direita em relação a Mizuno/Olympikus).
# Cada entrada é uma lista de possíveis grafias normalizadas do cabeçalho;
# a primeira que bater na linha 1 da aba é usada.
HEADER_ALIASES = {
    "sku": ["sku"],
    "codigo_barras": ["codigo de barras", "codigodebarras", "cod barras"],
    "material_vulca_tam": ["materialvulcatam"],
    "material_vulca": ["materialvulca"],
    "artigo": ["artigo"],
    "descricao_item": ["descricao do item"],
    "nome_grupo": ["nome do grupo"],
    "colorway": ["colorway description", "colorway descri"],  # aceita truncado
}

# Ordem e nomes das colunas na aba final. Cada tupla é (nome_da_coluna_final,
# chave_do_header_acima). "marca" e "segmento" são calculadas, não vêm direto
# de um header.
FINAL_COLUMNS = [
    ("SKU", "sku"),
    ("Códigodebarras", "codigo_barras"),
    ("MaterialVulcaTam", "material_vulca_tam"),
    ("MaterialVulca", "material_vulca"),
    ("Artigo", "artigo"),
    ("SKU", "sku"),
    ("Descrição do item", "descricao_item"),
    ("Nome do grupo", "nome_grupo"),
    ("Códigodebarras", "codigo_barras"),
    ("Colorway Description", "colorway"),
    ("MARCA", None),      # calculada: nome da aba/arquivo de origem
    ("SEGMENTO", None),   # calculada: a partir de "nome do grupo"
]

# Marcas que não têm aba própria: vivem dentro da aba de outra marca e são
# identificadas apenas pelo texto de "Nome do grupo". Ex.: OPANKA não tem
# estoque próprio, mas tem cadastro na base de embalagem, e vem dentro da aba
# Olympikus. Formato: (substring procurada em "Nome do grupo", MARCA, SEGMENTO
# fixo ou None para cair nas regras normais).
MARCA_EMBUTIDA_RULES = [
    ("opanka", "OPANKA", "CHINELOS"),
]

# Regras que dependem do texto exato de "Nome do grupo" e têm precedência
# sobre as regras genéricas de segmento abaixo. Ex.: a linha de roupas do
# Botafogo da Mizuno vem marcada como "Futebol Mizuno", mas é vestuário —
# a palavra "futebol" sozinha não diria isso.
SEGMENTO_OVERRIDES = [
    ("futebol mizuno", "VESTUÁRIOS"),
]

# Regras de tradução/padronização do SEGMENTO a partir do texto de
# "Nome do grupo" (comparação por substring, case/acento-insensitive).
# A ordem importa: a primeira regra que bater é usada — por isso "chinelos" e
# "meias" vêm antes de "calcados"/"vestuario", já que um grupo pode conter
# as duas palavras e o mais específico deve ganhar.
SEGMENTO_RULES = [
    ("chinelos", "CHINELOS"),
    ("chinelo", "CHINELOS"),
    ("sandals", "CHINELOS"),
    ("slides", "CHINELOS"),
    ("meias", "MEIAS"),
    ("meia", "MEIAS"),
    ("socks", "MEIAS"),
    ("chuteiras", "CALÇADOS"),
    ("chuteira", "CALÇADOS"),
    ("calcados", "CALÇADOS"),
    ("footwear", "CALÇADOS"),
    ("acessorios", "ACESSÓRIOS"),
    ("accessories", "ACESSÓRIOS"),
    ("apparel", "VESTUÁRIOS"),
    ("vestuario", "VESTUÁRIOS"),
    ("vestuarios", "VESTUÁRIOS"),
]

FINAL_SHEET_NAME = "Base Final"

# Nome padrão (sem extensão) do arquivo de saída quando --output não é
# informado. O arquivo sempre é salvo como .xlsx.
DEFAULT_OUTPUT_NAME = "Embalas"


# ---------------------------------------------------------------------------
# Camada de leitura: unifica .xlsx (openpyxl) e .xlsb (pyxlsb) atrás da mesma
# interface simples (nome das abas, linha de cabeçalho, linhas de dados).
# ---------------------------------------------------------------------------

class _XlsxSource:
    def __init__(self, path):
        self._wb = load_workbook(path, data_only=True, read_only=True)

    @property
    def sheetnames(self):
        return self._wb.sheetnames

    def header_row(self, sheet_name):
        ws = self._wb[sheet_name]
        return [c.value for c in next(ws.iter_rows(max_row=1))]

    def data_rows(self, sheet_name):
        ws = self._wb[sheet_name]
        yield from ws.iter_rows(min_row=2, values_only=True)


class _XlsbSource:
    """Lê .xlsb via pyxlsb. Linhas vêm como listas esparsas de células
    (r, c, v); reconstruímos uma lista densa por índice de coluna."""

    def __init__(self, path):
        if pyxlsb is None:
            raise RuntimeError(
                "Este arquivo é .xlsb, mas a biblioteca 'pyxlsb' não está instalada. "
                "Rode: pip install pyxlsb"
            )
        self._wb = pyxlsb.open_workbook(str(path))

    @property
    def sheetnames(self):
        return list(self._wb.sheets)

    def _dense_rows(self, sheet_name):
        with self._wb.get_sheet(sheet_name) as ws:
            for row in ws.rows():
                if not row:
                    yield []
                    continue
                width = max(cell.c for cell in row) + 1
                dense = [None] * width
                for cell in row:
                    dense[cell.c] = cell.v
                yield dense

    def header_row(self, sheet_name):
        return next(self._dense_rows(sheet_name))

    def data_rows(self, sheet_name):
        rows = self._dense_rows(sheet_name)
        next(rows)  # pula cabeçalho
        yield from rows


def open_source(path: Path):
    if path.suffix.lower() == ".xlsb":
        return _XlsbSource(path)
    return _XlsxSource(path)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _as_barcode_number(val):
    """
    Converte o código de barras para número inteiro puro (sem notação
    científica, sem casas decimais) — é assim que o usuário quer a coluna
    formatada na base final. Alguns exports trazem o valor como texto
    ('7894756748414') e outros como float (7894756748414.0); ambos viram int.

    Não mexe em valores vazios nem em códigos com zero à esquerda (ex.:
    '0894756748414') — converter esses para int apagaria o zero e mudaria o
    código de verdade, então esses ficam como estavam para revisão manual.
    """
    if val is None:
        return None
    if isinstance(val, float):
        return int(val) if val.is_integer() else val
    if isinstance(val, int):
        return val
    text = str(val).strip()
    if text.isdigit() and not (len(text) > 1 and text[0] == "0"):
        return int(text)
    return val


def _norm(text) -> str:
    """Normaliza texto para comparação: sem acento, minúsculo, sem espaços extras."""
    if text is None:
        return ""
    text = str(text).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return " ".join(text.split())


def _alias_match(name, aliases):
    norm_name = _norm(name)
    return any(_norm(a) in norm_name or norm_name in _norm(a) for a in aliases)


def locate_brand_sheets(sources, source_paths):
    """
    Para cada marca em SOURCE_SHEETS, encontra em qual (source, sheet_name) ela
    está — primeiro tentando casar o nome da ABA com os aliases da marca; se
    nenhuma aba bater em nenhum arquivo, cai para casar o nome do ARQUIVO
    (útil quando cada marca chega em um arquivo separado com aba genérica
    tipo "Plan1").

    Retorna {marca: (source, sheet_name) ou None}.
    """
    result = {}
    for marca, aliases in SOURCE_SHEETS.items():
        found = None
        for src, path in zip(sources, source_paths):
            for sheet_name in src.sheetnames:
                if _alias_match(sheet_name, aliases):
                    found = (src, sheet_name)
                    break
            if found:
                break
        if found is None:
            for src, path in zip(sources, source_paths):
                if _alias_match(path.stem, aliases) and len(src.sheetnames) == 1:
                    found = (src, src.sheetnames[0])
                    break
        result[marca] = found
    return result


def map_headers(header_row):
    """
    header_row: lista de valores da linha 1 da aba (na ordem das colunas).
    Retorna {header_key: column_index (0-based)} para os headers que foram encontrados.
    """
    # Correspondência EXATA (após normalizar acento/caixa) — importante para não
    # confundir "SKU" com uma coluna parecida como "SKUCOMERCIAL?".
    found = {}
    normalized = [_norm(v) for v in header_row]
    for key, aliases in HEADER_ALIASES.items():
        alias_norms = {_norm(a) for a in aliases}
        for col_idx, cell_norm in enumerate(normalized):
            if cell_norm in alias_norms:
                found[key] = col_idx
                break
    return found


def compute_marca_segmento(nome_grupo_valor, marca_da_aba):
    """
    Decide MARCA e SEGMENTO de uma linha a partir do texto de "Nome do grupo".

    A MARCA normalmente é a aba/arquivo de origem, mas algumas marcas não têm
    aba própria e só se distinguem pelo "Nome do grupo" (ver
    MARCA_EMBUTIDA_RULES), então essa checagem vem primeiro.

    Retorna (marca, segmento, segmento_foi_mapeado).
    """
    norm = _norm(nome_grupo_valor)

    # 1) Marca embutida na aba de outra marca (ex.: OPANKA dentro de Olympikus).
    for needle, marca, segmento_fixo in MARCA_EMBUTIDA_RULES:
        if needle in norm:
            if segmento_fixo:
                return marca, segmento_fixo, True
            marca_da_aba = marca
            break

    # 2) Overrides de segmento por grupo específico (ex.: Futebol Mizuno).
    for needle, segmento in SEGMENTO_OVERRIDES:
        if needle in norm:
            return marca_da_aba, segmento, True

    # 3) Regras genéricas de tradução/padronização.
    for needle, segmento in SEGMENTO_RULES:
        if needle in norm:
            return marca_da_aba, segmento, True

    # Nada bateu: mantém o texto original para revisão manual.
    original = str(nome_grupo_valor).strip() if nome_grupo_valor else ""
    return marca_da_aba, original, False


# ---------------------------------------------------------------------------
# Núcleo da consolidação
# ---------------------------------------------------------------------------

def consolidate(input_paths):
    sources = [open_source(p) for p in input_paths]
    located = locate_brand_sheets(sources, input_paths)

    report_lines = []
    all_rows = []  # linhas já no formato final (lista de valores)
    missing_field_counts = {}  # marca -> quantidade de linhas com campo faltando
    unmapped_segmentos = set()
    marca_counts = {}  # MARCA final -> quantidade de linhas (pode diferir da aba,
                       # ex.: OPANKA sai de dentro da aba Olympikus)

    for marca, found in located.items():
        if found is None:
            report_lines.append(f"⚠️  Marca '{marca}' não encontrada em nenhum arquivo — pulando.")
            continue
        src, sheet_name = found

        header_row = src.header_row(sheet_name)
        header_map = map_headers(header_row)

        missing_headers = [k for k in HEADER_ALIASES if k not in header_map]
        if missing_headers:
            report_lines.append(
                f"⚠️  Marca '{marca}' (aba '{sheet_name}'): cabeçalhos não encontrados: "
                f"{', '.join(missing_headers)} (essas colunas ficarão em branco na base final)."
            )

        rows_in_sheet = 0
        missing_in_sheet = 0

        for row in src.data_rows(sheet_name):
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue  # linha totalmente vazia

            def get(key):
                idx = header_map.get(key)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            nome_grupo_valor = get("nome_grupo")
            marca_linha, segmento, mapped_ok = compute_marca_segmento(nome_grupo_valor, marca)
            if not mapped_ok and nome_grupo_valor:
                unmapped_segmentos.add(str(nome_grupo_valor).strip())
            marca_counts[marca_linha] = marca_counts.get(marca_linha, 0) + 1

            final_row = []
            row_missing = False
            for col_name, header_key in FINAL_COLUMNS:
                if header_key == "sku":
                    val = get("sku")
                elif header_key == "codigo_barras":
                    val = _as_barcode_number(get("codigo_barras"))
                elif header_key is not None:
                    val = get(header_key)
                elif col_name == "MARCA":
                    val = marca_linha
                elif col_name == "SEGMENTO":
                    val = segmento
                else:
                    val = None
                if header_key is not None and (val is None or str(val).strip() == ""):
                    row_missing = True
                final_row.append(val)

            all_rows.append(final_row)
            rows_in_sheet += 1
            if row_missing:
                missing_in_sheet += 1

        missing_field_counts[marca] = missing_in_sheet
        report_lines.append(f"✓ Marca '{marca}' (aba '{sheet_name}'): {rows_in_sheet} linhas consolidadas.")

    if marca_counts:
        resumo_marcas = ", ".join(f"{m}: {c}" for m, c in sorted(marca_counts.items()))
        report_lines.append(f"Linhas por MARCA na base final — {resumo_marcas}")

    if unmapped_segmentos:
        report_lines.append(
            "⚠️  Valores de 'Nome do grupo' sem regra de SEGMENTO mapeada (mantidos como estavam, "
            "revisar manualmente): " + "; ".join(sorted(unmapped_segmentos))
        )

    for marca, count in missing_field_counts.items():
        if count:
            report_lines.append(f"⚠️  Marca '{marca}': {count} linha(s) com pelo menos um campo ausente.")

    return all_rows, report_lines


def write_output(rows, output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = FINAL_SHEET_NAME

    headers = [name for name, _ in FINAL_COLUMNS]
    ws.append(headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        ws.append(row)

    # Código de barras: número inteiro puro, sem notação científica e sem
    # separador de milhar (formato "0" no Excel) — pedido explícito do
    # usuário, não o "Geral" padrão que faz o Excel mostrar algo como
    # "7,89476E+12" em números grandes.
    barcode_cols = [i for i, (name, _) in enumerate(FINAL_COLUMNS, start=1) if name == "Códigodebarras"]
    for col_idx in barcode_cols:
        for row_idx in range(2, len(rows) + 2):
            ws.cell(row=row_idx, column=col_idx).number_format = "0"

    # Auto-largura simples baseada no maior valor de cada coluna (amostra as
    # primeiras linhas para não pesar em bases muito grandes).
    sample = rows[:2000]
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in sample:
            v = row[col_idx - 1]
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    ws.freeze_panes = "A2"
    wb.save(output_path)


def main():
    args = sys.argv[1:]
    output_path = None
    if "--output" in args:
        i = args.index("--output")
        output_path = Path(args[i + 1])
        del args[i:i + 2]

    if not args:
        print("Uso: python consolidate.py <arquivo1> [<arquivo2> <arquivo3> ...] [--output <arquivo_saida.xlsx>]")
        sys.exit(1)

    input_paths = [Path(a) for a in args]
    for p in input_paths:
        if not p.exists():
            print(f"Arquivo não encontrado: {p}")
            sys.exit(1)

    if output_path is None:
        output_path = input_paths[0].with_name(f"{DEFAULT_OUTPUT_NAME}.xlsx")

    rows, report_lines = consolidate(input_paths)
    write_output(rows, output_path)

    print(f"\nBase final gerada: {output_path}")
    print(f"Total de linhas consolidadas: {len(rows)}\n")
    print("Resumo:")
    for line in report_lines:
        print(f"  {line}")


if __name__ == "__main__":
    main()
